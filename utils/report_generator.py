import os
from datetime import datetime
from typing import List
from checkers.base_checker import CheckResult
from utils import crypto, config_loader


def generate(results: List[CheckResult], fmt: str, path: str, system_info: dict) -> None:
    fmt = fmt.lower()
    if fmt == "html":
        content = _generate_html(results, system_info)
        _write_text(path, content)
    elif fmt == "txt":
        content = _generate_txt(results, system_info)
        _write_text(path, content)
    elif fmt == "excel":
        _generate_excel(results, system_info, path)
        content = _generate_txt(results, system_info)
    elif fmt == "pdf":
        content = _generate_html(results, system_info)
        _generate_pdf(content, path)
        content = _generate_txt(results, system_info)
    else:
        raise ValueError(f"Unsupported format: {fmt}")

    enc_path = os.path.splitext(path)[0] + "_encrypted.sec"
    password = config_loader.get("report_password", "Sec@2026")
    plain = content if fmt in ("html", "txt") else _generate_txt(results, system_info)
    encrypted = crypto.encrypt(plain, password)
    _write_text(enc_path, encrypted)


def _write_text(path: str, content: str) -> None:
    with open(path, "w", encoding="utf-8") as f:
        f.write(content)


def _generate_txt(results: List[CheckResult], system_info: dict) -> str:
    lines = []
    lines.append("=" * 60)
    lines.append("         终端安全合规检查报告")
    lines.append("=" * 60)
    lines.append(f"设备名称: {system_info.get('hostname', 'N/A')}")
    lines.append(f"操作系统: {system_info.get('os', 'N/A')}")
    lines.append(f"MAC地址:  {system_info.get('mac', 'N/A')}")
    lines.append(f"IP地址:   {system_info.get('ip', 'N/A')}")
    lines.append(f"检查时间: {system_info.get('check_time', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))}")
    lines.append("=" * 60)

    for r in results:
        lines.append(f"\n{'─' * 40}")
        status = "【合规】" if r.passed else "【不合规】"
        lines.append(f"{status} {r.module}")
        lines.append("─" * 40)
        lines.append(r.summary)
        if r.recommendations:
            lines.append("\n整改建议:")
            for rec in r.recommendations:
                lines.append(f"  · {rec}")

    lines.append("\n" + "=" * 60)
    return "\n".join(lines)


def _generate_html(results: List[CheckResult], system_info: dict) -> str:
    rows = ""
    for r in results:
        status_class = "compliant" if r.passed else "non-compliant"
        status_text = "合规" if r.passed else "不合规"
        summary_html = r.summary.replace("\n", "<br>").replace(
            "✗ 不合规", '<span class="violation">✗ 不合规</span>'
        )
        recs = "".join(f"<li>{rec}</li>" for rec in r.recommendations)
        recs_html = f"<ul>{recs}</ul>" if recs else ""
        rows += f"""
        <div class="module">
          <h2><span class="badge {status_class}">{status_text}</span> {r.module}</h2>
          <pre>{summary_html}</pre>
          {recs_html}
        </div>"""

    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<title>终端安全合规检查报告</title>
<style>
  body {{ font-family: "Microsoft YaHei", "SimSun", sans-serif; margin: 40px; background: #f5f5f5; color: #333; }}
  h1 {{ text-align: center; color: #1a1a2e; }}
  .info-box {{ background: #fff; border-left: 4px solid #0066cc; padding: 16px 20px; margin-bottom: 24px; border-radius: 4px; }}
  .info-box p {{ margin: 4px 0; }}
  .module {{ background: #fff; padding: 16px 20px; margin-bottom: 16px; border-radius: 4px; box-shadow: 0 1px 3px rgba(0,0,0,.1); }}
  h2 {{ margin-top: 0; font-size: 1.1em; }}
  pre {{ background: #f8f8f8; padding: 12px; border-radius: 4px; white-space: pre-wrap; font-size: 0.9em; }}
  .badge {{ padding: 2px 8px; border-radius: 4px; font-size: 0.85em; margin-right: 8px; }}
  .compliant {{ background: #d4edda; color: #155724; }}
  .non-compliant {{ background: #f8d7da; color: #721c24; }}
  .violation {{ color: #dc3545; font-weight: bold; }}
  ul {{ margin: 8px 0; padding-left: 20px; }}
  li {{ margin: 4px 0; color: #856404; }}
</style>
</head>
<body>
<h1>终端安全合规检查报告</h1>
<div class="info-box">
  <p><strong>设备名称：</strong>{system_info.get('hostname', 'N/A')}</p>
  <p><strong>操作系统：</strong>{system_info.get('os', 'N/A')}</p>
  <p><strong>MAC地址：</strong>{system_info.get('mac', 'N/A')}</p>
  <p><strong>IP地址：</strong>{system_info.get('ip', 'N/A')}</p>
  <p><strong>检查时间：</strong>{system_info.get('check_time', '')}</p>
</div>
{rows}
</body>
</html>"""


def _generate_excel(results: List[CheckResult], system_info: dict, path: str) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    info_ws = wb.create_sheet("设备信息")
    info_ws.append(["字段", "值"])
    for k, v in system_info.items():
        info_ws.append([k, v])

    for r in results:
        ws = wb.create_sheet(r.module[:30])
        ws.append(["状态", "合规" if r.passed else "不合规"])
        ws.append(["摘要", r.summary])
        ws.append([])
        ws.append(["违规详情"])
        for v in r.violations:
            ws.append([str(v)])
        ws.append([])
        ws.append(["整改建议"])
        for rec in r.recommendations:
            ws.append([rec])

        red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        ws["A1"].fill = green_fill if r.passed else red_fill

    wb.save(path)


def _generate_pdf(html_content: str, path: str) -> None:
    try:
        from PyQt5.QtWidgets import QApplication
        from PyQt5.QtPrintSupport import QPrinter
        from PyQt5.QtGui import QTextDocument
        import sys

        app = QApplication.instance() or QApplication(sys.argv)
        printer = QPrinter(QPrinter.HighResolution)
        printer.setOutputFormat(QPrinter.PdfFormat)
        printer.setOutputFileName(path)

        doc = QTextDocument()
        doc.setHtml(html_content)
        doc.print_(printer)
    except Exception as e:
        txt_path = os.path.splitext(path)[0] + ".txt"
        raise RuntimeError(f"PDF generation failed, save as TXT instead ({txt_path}): {e}")
